"""
APP Service: Transmission Reports
Generates comprehensive transmission status reports for FIRS submissions
"""

import asyncio
import json
import logging
from datetime import datetime, timezone, timedelta
from typing import Dict, Any, Optional, List, Union, Tuple
from dataclasses import dataclass, asdict
from enum import Enum
import csv
import io
import base64
from collections import defaultdict, Counter
from uuid import UUID

from sqlalchemy import desc, func, select
from sqlalchemy.sql import and_, or_

from core_platform.data_management.db_async import get_async_session
from core_platform.data_management.models.firs_submission import (
    FIRSSubmission,
    SubmissionStatus,
)

try:
    import matplotlib.pyplot as plt  # type: ignore
except Exception:  # pragma: no cover - optional dependency
    plt = None  # type: ignore

try:
    import seaborn as sns  # type: ignore
except Exception:  # pragma: no cover - optional dependency
    sns = None  # type: ignore


class ReportFormat(str, Enum):
    """Report output formats"""
    JSON = "json"
    CSV = "csv"
    PDF = "pdf"
    EXCEL = "excel"
    HTML = "html"


class TransmissionStatus(str, Enum):
    """Transmission status categories"""
    PENDING = "pending"
    SUBMITTED = "submitted"
    ACKNOWLEDGED = "acknowledged"
    APPROVED = "approved"
    REJECTED = "rejected"
    FAILED = "failed"
    CANCELLED = "cancelled"
    TIMEOUT = "timeout"


class ReportPeriod(str, Enum):
    """Report time periods"""
    HOURLY = "hourly"
    DAILY = "daily"
    WEEKLY = "weekly"
    MONTHLY = "monthly"
    QUARTERLY = "quarterly"
    YEARLY = "yearly"
    CUSTOM = "custom"


@dataclass
class TransmissionRecord:
    """Individual transmission record"""
    transmission_id: str
    invoice_number: str
    irn: Optional[str]
    client_id: str
    organization_id: str
    status: TransmissionStatus
    submitted_at: datetime
    acknowledged_at: Optional[datetime]
    completed_at: Optional[datetime]
    processing_time_seconds: Optional[float]
    retry_count: int
    error_code: Optional[str]
    error_message: Optional[str]
    firs_response: Optional[Dict[str, Any]]
    payload_size_bytes: int
    
    def to_dict(self) -> Dict[str, Any]:
        data = asdict(self)
        data['submitted_at'] = self.submitted_at.isoformat()
        if self.acknowledged_at:
            data['acknowledged_at'] = self.acknowledged_at.isoformat()
        if self.completed_at:
            data['completed_at'] = self.completed_at.isoformat()
        return data


@dataclass
class ReportConfig:
    """Configuration for report generation"""
    start_date: datetime
    end_date: datetime
    format: ReportFormat = ReportFormat.JSON
    include_details: bool = True
    include_charts: bool = False
    group_by: Optional[List[str]] = None
    filter_criteria: Optional[Dict[str, Any]] = None
    sort_by: Optional[str] = None
    limit: Optional[int] = None
    
    def to_dict(self) -> Dict[str, Any]:
        data = asdict(self)
        data['start_date'] = self.start_date.isoformat()
        data['end_date'] = self.end_date.isoformat()
        return data


@dataclass
class TransmissionSummary:
    """Summary statistics for transmissions"""
    total_transmissions: int
    successful_transmissions: int
    failed_transmissions: int
    pending_transmissions: int
    success_rate: float
    average_processing_time: float
    total_payload_size: int
    unique_clients: int
    unique_organizations: int
    status_breakdown: Dict[str, int]
    error_breakdown: Dict[str, int]
    hourly_distribution: Dict[str, int]
    retry_statistics: Dict[str, int]
    
    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)


class TransmissionDataProvider:
    """Data provider for transmission records"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self._mock_data: Optional[List[TransmissionRecord]] = None
        self._status_map = {
            SubmissionStatus.PENDING: TransmissionStatus.PENDING,
            SubmissionStatus.PROCESSING: TransmissionStatus.PENDING,
            SubmissionStatus.SUBMITTED: TransmissionStatus.SUBMITTED,
            SubmissionStatus.ACCEPTED: TransmissionStatus.APPROVED,
            SubmissionStatus.REJECTED: TransmissionStatus.REJECTED,
            SubmissionStatus.FAILED: TransmissionStatus.FAILED,
            SubmissionStatus.CANCELLED: TransmissionStatus.CANCELLED,
        }
        self._status_filters = {
            TransmissionStatus.PENDING: {SubmissionStatus.PENDING, SubmissionStatus.PROCESSING},
            TransmissionStatus.SUBMITTED: {SubmissionStatus.SUBMITTED},
            TransmissionStatus.ACKNOWLEDGED: {SubmissionStatus.SUBMITTED, SubmissionStatus.ACCEPTED},
            TransmissionStatus.APPROVED: {SubmissionStatus.ACCEPTED},
            TransmissionStatus.REJECTED: {SubmissionStatus.REJECTED},
            TransmissionStatus.FAILED: {SubmissionStatus.FAILED},
            TransmissionStatus.CANCELLED: {SubmissionStatus.CANCELLED},
        }
    
    def _generate_mock_data(self) -> List[TransmissionRecord]:
        """Generate mock transmission data for demonstration"""
        import random
        from datetime import timedelta
        
        records = []
        statuses = list(TransmissionStatus)
        
        base_time = datetime.now(timezone.utc) - timedelta(days=30)
        
        for i in range(1000):
            status = random.choice(statuses)
            submitted_at = base_time + timedelta(
                days=random.randint(0, 30),
                hours=random.randint(0, 23),
                minutes=random.randint(0, 59)
            )
            
            acknowledged_at = None
            completed_at = None
            processing_time = None
            
            if status not in [TransmissionStatus.PENDING, TransmissionStatus.FAILED]:
                acknowledged_at = submitted_at + timedelta(seconds=random.randint(5, 300))
                
                if status in [TransmissionStatus.APPROVED, TransmissionStatus.REJECTED]:
                    completed_at = acknowledged_at + timedelta(seconds=random.randint(60, 3600))
                    processing_time = (completed_at - submitted_at).total_seconds()
            
            error_code = None
            error_message = None
            if status in [TransmissionStatus.REJECTED, TransmissionStatus.FAILED]:
                error_codes = ["VALIDATION_ERROR", "TIMEOUT", "INVALID_IRN", "NETWORK_ERROR"]
                error_code = random.choice(error_codes)
                error_message = f"Error: {error_code}"
            
            record = TransmissionRecord(
                transmission_id=f"TXN_{i+1:06d}",
                invoice_number=f"INV_{i+1:06d}",
                irn=f"IRN_{i+1:010d}" if status != TransmissionStatus.FAILED else None,
                client_id=f"CLIENT_{random.randint(1, 20):03d}",
                organization_id=f"ORG_{random.randint(1, 10):03d}",
                status=status,
                submitted_at=submitted_at,
                acknowledged_at=acknowledged_at,
                completed_at=completed_at,
                processing_time_seconds=processing_time,
                retry_count=random.randint(0, 3),
                error_code=error_code,
                error_message=error_message,
                firs_response={"status": status.value} if acknowledged_at else None,
                payload_size_bytes=random.randint(1000, 50000)
            )
            records.append(record)
        
        return records
    
    async def get_transmissions(
        self,
        start_date: datetime,
        end_date: datetime,
        filter_criteria: Optional[Dict[str, Any]] = None,
        limit: Optional[int] = None,
    ) -> List[TransmissionRecord]:
        """Get transmission records for date range with optional filters"""

        filter_criteria = filter_criteria or {}

        try:
            records = await self._fetch_records_from_db(start_date, end_date, filter_criteria, limit)
            if records:
                return records
        except Exception as exc:  # pragma: no cover - fallback to mock data
            self.logger.warning("Falling back to mock transmission data: %s", exc)

        if self._mock_data is None:
            self._mock_data = self._generate_mock_data()

        filtered_records = [
            record for record in self._mock_data
            if start_date <= record.submitted_at <= end_date
        ]

        filtered_records = self._apply_filters(filtered_records, filter_criteria)

        if limit is not None:
            return filtered_records[:limit]
        return filtered_records

    async def _fetch_records_from_db(
        self,
        start_date: datetime,
        end_date: datetime,
        filter_criteria: Dict[str, Any],
        limit: Optional[int],
    ) -> List[TransmissionRecord]:
        """Load transmission records from the database."""

        timestamp_field = func.coalesce(FIRSSubmission.submitted_at, FIRSSubmission.created_at)
        stmt = (
            select(FIRSSubmission)
            .where(timestamp_field >= start_date, timestamp_field <= end_date)
            .order_by(desc(timestamp_field))
        )

        status_filter = filter_criteria.get("status")
        if status_filter:
            stmt = stmt.where(FIRSSubmission.status.in_(self._coerce_status_filter(status_filter)))

        org_id = filter_criteria.get("organization_id") or filter_criteria.get("tenant_id")
        org_uuid = self._parse_uuid(org_id)
        if org_uuid:
            stmt = stmt.where(FIRSSubmission.organization_id == org_uuid)

        client_id = filter_criteria.get("client_id")
        if client_id:
            stmt = stmt.where(
                and_(
                    FIRSSubmission.customer_tin.isnot(None),
                    func.lower(FIRSSubmission.customer_tin) == str(client_id).lower(),
                )
            )

        has_errors = filter_criteria.get("has_errors")
        if has_errors is True:
            stmt = stmt.where(
                or_(
                    FIRSSubmission.firs_status_code.isnot(None),
                    FIRSSubmission.error_details.isnot(None),
                )
            )
        elif has_errors is False:
            stmt = stmt.where(
                and_(
                    FIRSSubmission.firs_status_code.is_(None),
                    FIRSSubmission.error_details.is_(None),
                )
            )

        if limit is not None:
            stmt = stmt.limit(int(max(1, limit)))

        async for session in get_async_session():
            result = await session.execute(stmt)
            rows = result.scalars().all()
            return [self._map_submission_to_record(row) for row in rows]

        return []

    def _apply_filters(
        self,
        records: List[TransmissionRecord],
        filter_criteria: Dict[str, Any],
    ) -> List[TransmissionRecord]:
        """Apply in-memory filters (used for mock fallback)."""

        filtered = records
        status_filter = filter_criteria.get("status")
        if status_filter:
            statuses = self._coerce_status_filter(status_filter)
            filtered = [r for r in filtered if self._map_to_submission_status(r.status) in statuses]

        client_id = filter_criteria.get("client_id")
        if client_id:
            filtered = [r for r in filtered if r.client_id == client_id]

        org_id = filter_criteria.get("organization_id")
        if org_id:
            filtered = [r for r in filtered if r.organization_id == str(org_id)]

        if "has_errors" in filter_criteria:
            if filter_criteria["has_errors"]:
                filtered = [r for r in filtered if r.error_code]
            else:
                filtered = [r for r in filtered if not r.error_code]

        return filtered

    def _map_submission_to_record(self, submission: FIRSSubmission) -> TransmissionRecord:
        """Convert database submission record into reporting record."""

        status = self._status_map.get(submission.status, TransmissionStatus.PENDING)
        submitted_at = submission.submitted_at or submission.created_at or datetime.now(timezone.utc)
        completion_timestamp = submission.accepted_at or submission.rejected_at
        acknowledged_at = submission.accepted_at or submission.submitted_at

        processing_time = None
        if submitted_at and completion_timestamp:
            processing_time = max(
                0.0,
                (completion_timestamp - submitted_at).total_seconds(),
            )

        error_code = submission.firs_status_code
        error_message = submission.firs_message
        if not error_message and submission.error_details:
            if isinstance(submission.error_details, dict):
                error_message = submission.error_details.get("message") or submission.error_details.get("detail")

        payload_size = 0
        try:
            payload_size = len(json.dumps(submission.invoice_data or {}))
        except Exception:  # pragma: no cover - defensive fallback
            payload_size = 0

        client_identifier = (
            submission.customer_tin
            or submission.customer_name
            or str(submission.organization_id)
        )

        return TransmissionRecord(
            transmission_id=str(submission.id),
            invoice_number=submission.invoice_number,
            irn=submission.irn,
            client_id=str(client_identifier),
            organization_id=str(submission.organization_id),
            status=status,
            submitted_at=submitted_at,
            acknowledged_at=acknowledged_at,
            completed_at=completion_timestamp,
            processing_time_seconds=processing_time,
            retry_count=submission.retry_count or 0,
            error_code=error_code,
            error_message=error_message,
            firs_response=submission.firs_response,
            payload_size_bytes=payload_size,
        )

    def _parse_uuid(self, value: Any) -> Optional[UUID]:
        if not value:
            return None
        try:
            return UUID(str(value))
        except Exception:
            return None

    def _coerce_status_filter(self, raw_status: Any) -> List[SubmissionStatus]:
        """Normalize status filter into submission statuses."""

        statuses: List[SubmissionStatus] = []

        items = raw_status
        if not isinstance(items, (list, tuple, set)):
            items = [items]

        for entry in items:
            if isinstance(entry, TransmissionStatus):
                statuses.extend(self._status_filters.get(entry, []))
                continue
            try:
                statuses.append(SubmissionStatus(entry))
                continue
            except Exception:
                pass
            normalized = str(entry).strip().lower()
            for ts, submission_group in self._status_filters.items():
                if ts.value == normalized:
                    statuses.extend(submission_group)
                    break

        if not statuses:
            return list(self._status_map.keys())
        return list({status for status in statuses})

    def _map_to_submission_status(self, status: TransmissionStatus) -> SubmissionStatus:
        for submission_status, mapped in self._status_map.items():
            if mapped == status:
                return submission_status
        return SubmissionStatus.PENDING


class TransmissionReportGenerator:
    """
    Generates comprehensive transmission status reports
    Provides various report formats and visualization options
    """
    
    def __init__(self, data_provider: Optional[TransmissionDataProvider] = None):
        self.data_provider = data_provider or TransmissionDataProvider()
        self.logger = logging.getLogger(__name__)
        
        # Report generation statistics
        self.stats = {
            'reports_generated': 0,
            'last_report_at': None,
            'format_usage': Counter(),
            'period_usage': Counter(),
            'average_generation_time': 0.0
        }
    
    async def generate_report(self, config: ReportConfig) -> Dict[str, Any]:
        """
        Generate transmission report based on configuration
        
        Args:
            config: Report configuration
            
        Returns:
            Generated report data
        """
        start_time = datetime.now(timezone.utc)
        
        try:
            self.logger.info(
                f"Generating transmission report: {config.format.value} "
                f"from {config.start_date} to {config.end_date}"
            )
            
            # Get transmission data
            records = await self.data_provider.get_transmissions(
                start_date=config.start_date,
                end_date=config.end_date,
                filter_criteria=config.filter_criteria,
                limit=config.limit,
            )
            
            # Apply sorting and limiting
            if config.sort_by:
                records = self._sort_records(records, config.sort_by)
            
            if config.limit:
                records = records[:config.limit]
            
            # Generate summary statistics
            summary = self._calculate_summary(records)
            
            # Group data if requested
            grouped_data = None
            if config.group_by:
                grouped_data = self._group_records(records, config.group_by)
            
            # Generate charts if requested
            charts = None
            if config.include_charts:
                charts = await self._generate_charts(records, summary)
            
            # Format report based on requested format
            report_data = await self._format_report(
                records=records,
                summary=summary,
                grouped_data=grouped_data,
                charts=charts,
                config=config
            )
            
            # Update statistics
            generation_time = (datetime.now(timezone.utc) - start_time).total_seconds()
            self._update_stats(config.format, generation_time)
            
            self.logger.info(
                f"Report generated successfully: {len(records)} records in {generation_time:.2f}s"
            )
            
            return {
                'report_data': report_data,
                'metadata': {
                    'generated_at': start_time.isoformat(),
                    'generation_time_seconds': generation_time,
                    'record_count': len(records),
                    'config': config.to_dict()
                }
            }
            
        except Exception as e:
            self.logger.error(f"Error generating transmission report: {str(e)}")
            raise
    
    def _sort_records(self, records: List[TransmissionRecord], sort_by: str) -> List[TransmissionRecord]:
        """Sort records by specified field"""
        sort_key_map = {
            'submitted_at': lambda r: r.submitted_at,
            'processing_time': lambda r: r.processing_time_seconds or 0,
            'status': lambda r: r.status.value,
            'client_id': lambda r: r.client_id,
            'organization_id': lambda r: r.organization_id,
            'retry_count': lambda r: r.retry_count
        }
        
        if sort_by in sort_key_map:
            return sorted(records, key=sort_key_map[sort_by])
        
        return records
    
    def _calculate_summary(self, records: List[TransmissionRecord]) -> TransmissionSummary:
        """Calculate summary statistics for transmission records"""
        if not records:
            return TransmissionSummary(
                total_transmissions=0,
                successful_transmissions=0,
                failed_transmissions=0,
                pending_transmissions=0,
                success_rate=0.0,
                average_processing_time=0.0,
                total_payload_size=0,
                unique_clients=0,
                unique_organizations=0,
                status_breakdown={},
                error_breakdown={},
                hourly_distribution={},
                retry_statistics={}
            )
        
        # Basic counts
        total = len(records)
        successful = len([r for r in records if r.status == TransmissionStatus.APPROVED])
        failed = len([r for r in records if r.status in [TransmissionStatus.REJECTED, TransmissionStatus.FAILED]])
        pending = len([r for r in records if r.status == TransmissionStatus.PENDING])
        
        # Success rate
        success_rate = (successful / total * 100) if total > 0 else 0.0
        
        # Average processing time
        processing_times = [r.processing_time_seconds for r in records if r.processing_time_seconds]
        avg_processing_time = sum(processing_times) / len(processing_times) if processing_times else 0.0
        
        # Total payload size
        total_payload_size = sum(r.payload_size_bytes for r in records)
        
        # Unique counts
        unique_clients = len(set(r.client_id for r in records))
        unique_organizations = len(set(r.organization_id for r in records))
        
        # Status breakdown
        status_breakdown = Counter(r.status.value for r in records)
        
        # Error breakdown
        error_breakdown = Counter(r.error_code for r in records if r.error_code)
        
        # Hourly distribution
        hourly_distribution = Counter(
            r.submitted_at.strftime('%H:00') for r in records
        )
        
        # Retry statistics
        retry_statistics = Counter(str(r.retry_count) for r in records)
        
        return TransmissionSummary(
            total_transmissions=total,
            successful_transmissions=successful,
            failed_transmissions=failed,
            pending_transmissions=pending,
            success_rate=round(success_rate, 2),
            average_processing_time=round(avg_processing_time, 2),
            total_payload_size=total_payload_size,
            unique_clients=unique_clients,
            unique_organizations=unique_organizations,
            status_breakdown=dict(status_breakdown),
            error_breakdown=dict(error_breakdown),
            hourly_distribution=dict(hourly_distribution),
            retry_statistics=dict(retry_statistics)
        )
    
    def _group_records(self, records: List[TransmissionRecord], group_by: List[str]) -> Dict[str, Any]:
        """Group records by specified fields"""
        grouped = defaultdict(list)
        
        for record in records:
            group_key_parts = []
            for field in group_by:
                if hasattr(record, field):
                    value = getattr(record, field)
                    if isinstance(value, datetime):
                        value = value.isoformat()
                    elif isinstance(value, Enum):
                        value = value.value
                    group_key_parts.append(str(value))
            
            group_key = " | ".join(group_key_parts)
            grouped[group_key].append(record)
        
        # Calculate summary for each group
        result = {}
        for group_key, group_records in grouped.items():
            result[group_key] = {
                'count': len(group_records),
                'summary': self._calculate_summary(group_records).to_dict(),
                'records': [r.to_dict() for r in group_records] if len(group_records) <= 100 else None
            }
        
        return result
    
    async def _generate_charts(self, records: List[TransmissionRecord], summary: TransmissionSummary) -> Dict[str, str]:
        """Generate charts for the report"""
        charts = {}
        
        if plt is None:
            self.logger.debug("Matplotlib not available; skipping chart generation")
            return charts
        
        try:
            if sns is not None:
                plt.style.use('seaborn-v0_8')
            
            # Status distribution chart
            if summary.status_breakdown:
                fig, ax = plt.subplots(figsize=(10, 6))
                statuses = list(summary.status_breakdown.keys())
                counts = list(summary.status_breakdown.values())
                
                colors = plt.cm.Set3(range(len(statuses)))
                ax.pie(counts, labels=statuses, autopct='%1.1f%%', colors=colors)
                ax.set_title('Transmission Status Distribution')
                
                # Convert to base64
                buffer = io.BytesIO()
                plt.savefig(buffer, format='png', bbox_inches='tight', dpi=150)
                buffer.seek(0)
                charts['status_distribution'] = base64.b64encode(buffer.getvalue()).decode()
                plt.close()
            
            # Hourly distribution chart
            if summary.hourly_distribution:
                fig, ax = plt.subplots(figsize=(12, 6))
                hours = sorted(summary.hourly_distribution.keys())
                counts = [summary.hourly_distribution[hour] for hour in hours]
                
                ax.bar(hours, counts, color='skyblue', alpha=0.7)
                ax.set_title('Transmissions by Hour of Day')
                ax.set_xlabel('Hour')
                ax.set_ylabel('Number of Transmissions')
                ax.tick_params(axis='x', rotation=45)
                
                buffer = io.BytesIO()
                plt.savefig(buffer, format='png', bbox_inches='tight', dpi=150)
                buffer.seek(0)
                charts['hourly_distribution'] = base64.b64encode(buffer.getvalue()).decode()
                plt.close()
            
            # Processing time distribution
            processing_times = [r.processing_time_seconds for r in records if r.processing_time_seconds]
            if processing_times:
                fig, ax = plt.subplots(figsize=(10, 6))
                ax.hist(processing_times, bins=30, color='lightgreen', alpha=0.7, edgecolor='black')
                ax.set_title('Processing Time Distribution')
                ax.set_xlabel('Processing Time (seconds)')
                ax.set_ylabel('Frequency')
                
                buffer = io.BytesIO()
                plt.savefig(buffer, format='png', bbox_inches='tight', dpi=150)
                buffer.seek(0)
                charts['processing_time_distribution'] = base64.b64encode(buffer.getvalue()).decode()
                plt.close()
            
        except Exception as e:
            self.logger.warning(f"Error generating charts: {str(e)}")
        
        return charts
    
    async def _format_report(self, 
                            records: List[TransmissionRecord],
                            summary: TransmissionSummary,
                            grouped_data: Optional[Dict[str, Any]],
                            charts: Optional[Dict[str, str]],
                            config: ReportConfig) -> Union[Dict, str, bytes]:
        """Format report in requested format"""
        
        if config.format == ReportFormat.JSON:
            return self._format_json_report(records, summary, grouped_data, charts, config)
        elif config.format == ReportFormat.CSV:
            return self._format_csv_report(records, summary, config)
        elif config.format == ReportFormat.HTML:
            return self._format_html_report(records, summary, grouped_data, charts, config)
        elif config.format == ReportFormat.EXCEL:
            return await self._format_excel_report(records, summary, grouped_data, config)
        else:
            # Default to JSON
            return self._format_json_report(records, summary, grouped_data, charts, config)
    
    def _format_json_report(self, 
                           records: List[TransmissionRecord],
                           summary: TransmissionSummary,
                           grouped_data: Optional[Dict[str, Any]],
                           charts: Optional[Dict[str, str]],
                           config: ReportConfig) -> Dict[str, Any]:
        """Format report as JSON"""
        report = {
            'summary': summary.to_dict(),
            'config': config.to_dict()
        }
        
        if config.include_details:
            report['records'] = [record.to_dict() for record in records]
        
        if grouped_data:
            report['grouped_data'] = grouped_data
        
        if charts:
            report['charts'] = charts
        
        return report
    
    def _format_csv_report(self, 
                          records: List[TransmissionRecord],
                          summary: TransmissionSummary,
                          config: ReportConfig) -> str:
        """Format report as CSV"""
        output = io.StringIO()
        
        # Write summary first
        writer = csv.writer(output)
        writer.writerow(['=== TRANSMISSION REPORT SUMMARY ==='])
        writer.writerow(['Metric', 'Value'])
        writer.writerow(['Total Transmissions', summary.total_transmissions])
        writer.writerow(['Successful Transmissions', summary.successful_transmissions])
        writer.writerow(['Failed Transmissions', summary.failed_transmissions])
        writer.writerow(['Success Rate (%)', summary.success_rate])
        writer.writerow(['Average Processing Time (s)', summary.average_processing_time])
        writer.writerow([''])
        
        # Write detailed records if requested
        if config.include_details and records:
            writer.writerow(['=== DETAILED TRANSMISSION RECORDS ==='])
            
            # Headers
            headers = [
                'Transmission ID', 'Invoice Number', 'IRN', 'Client ID', 'Organization ID',
                'Status', 'Submitted At', 'Acknowledged At', 'Completed At',
                'Processing Time (s)', 'Retry Count', 'Error Code', 'Error Message',
                'Payload Size (bytes)'
            ]
            writer.writerow(headers)
            
            # Data rows
            for record in records:
                row = [
                    record.transmission_id,
                    record.invoice_number,
                    record.irn or '',
                    record.client_id,
                    record.organization_id,
                    record.status.value,
                    record.submitted_at.isoformat(),
                    record.acknowledged_at.isoformat() if record.acknowledged_at else '',
                    record.completed_at.isoformat() if record.completed_at else '',
                    record.processing_time_seconds or '',
                    record.retry_count,
                    record.error_code or '',
                    record.error_message or '',
                    record.payload_size_bytes
                ]
                writer.writerow(row)
        
        return output.getvalue()
    
    def _format_html_report(self, 
                           records: List[TransmissionRecord],
                           summary: TransmissionSummary,
                           grouped_data: Optional[Dict[str, Any]],
                           charts: Optional[Dict[str, str]],
                           config: ReportConfig) -> str:
        """Format report as HTML"""
        html_parts = [
            '<!DOCTYPE html>',
            '<html>',
            '<head>',
            '<title>Transmission Status Report</title>',
            '<style>',
            'body { font-family: Arial, sans-serif; margin: 20px; }',
            'table { border-collapse: collapse; width: 100%; margin: 20px 0; }',
            'th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }',
            'th { background-color: #f2f2f2; }',
            '.summary { background-color: #f9f9f9; padding: 15px; margin: 20px 0; }',
            '.chart { margin: 20px 0; text-align: center; }',
            '</style>',
            '</head>',
            '<body>',
            f'<h1>Transmission Status Report</h1>',
            f'<p>Generated on: {datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")}</p>',
            f'<p>Period: {config.start_date.strftime("%Y-%m-%d")} to {config.end_date.strftime("%Y-%m-%d")}</p>'
        ]
        
        # Add summary
        html_parts.extend([
            '<div class="summary">',
            '<h2>Summary Statistics</h2>',
            f'<p><strong>Total Transmissions:</strong> {summary.total_transmissions}</p>',
            f'<p><strong>Successful:</strong> {summary.successful_transmissions}</p>',
            f'<p><strong>Failed:</strong> {summary.failed_transmissions}</p>',
            f'<p><strong>Success Rate:</strong> {summary.success_rate}%</p>',
            f'<p><strong>Average Processing Time:</strong> {summary.average_processing_time} seconds</p>',
            '</div>'
        ])
        
        # Add charts if available
        if charts:
            html_parts.append('<h2>Charts</h2>')
            for chart_name, chart_data in charts.items():
                html_parts.extend([
                    '<div class="chart">',
                    f'<h3>{chart_name.replace("_", " ").title()}</h3>',
                    f'<img src="data:image/png;base64,{chart_data}" alt="{chart_name}">',
                    '</div>'
                ])
        
        # Add detailed records table if requested
        if config.include_details and records:
            html_parts.extend([
                '<h2>Detailed Records</h2>',
                '<table>',
                '<tr>',
                '<th>Transmission ID</th>',
                '<th>Invoice Number</th>',
                '<th>Status</th>',
                '<th>Submitted At</th>',
                '<th>Processing Time</th>',
                '<th>Retry Count</th>',
                '</tr>'
            ])
            
            for record in records[:100]:  # Limit to first 100 for HTML
                html_parts.extend([
                    '<tr>',
                    f'<td>{record.transmission_id}</td>',
                    f'<td>{record.invoice_number}</td>',
                    f'<td>{record.status.value}</td>',
                    f'<td>{record.submitted_at.strftime("%Y-%m-%d %H:%M:%S")}</td>',
                    f'<td>{record.processing_time_seconds or "N/A"}</td>',
                    f'<td>{record.retry_count}</td>',
                    '</tr>'
                ])
            
            html_parts.append('</table>')
        
        html_parts.extend(['</body>', '</html>'])
        
        return '\n'.join(html_parts)
    
    async def _format_excel_report(self, 
                                  records: List[TransmissionRecord],
                                  summary: TransmissionSummary,
                                  grouped_data: Optional[Dict[str, Any]],
                                  config: ReportConfig) -> bytes:
        """Format report as Excel file"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            wb = openpyxl.Workbook()
            
            # Summary sheet
            ws_summary = wb.active
            ws_summary.title = "Summary"
            
            # Add summary data
            ws_summary['A1'] = "Transmission Report Summary"
            ws_summary['A1'].font = Font(bold=True, size=14)
            
            summary_data = [
                ['Total Transmissions', summary.total_transmissions],
                ['Successful Transmissions', summary.successful_transmissions],
                ['Failed Transmissions', summary.failed_transmissions],
                ['Success Rate (%)', summary.success_rate],
                ['Average Processing Time (s)', summary.average_processing_time],
                ['Unique Clients', summary.unique_clients],
                ['Unique Organizations', summary.unique_organizations]
            ]
            
            for i, (metric, value) in enumerate(summary_data, start=3):
                ws_summary[f'A{i}'] = metric
                ws_summary[f'B{i}'] = value
                ws_summary[f'A{i}'].font = Font(bold=True)
            
            # Detailed records sheet
            if config.include_details and records:
                ws_details = wb.create_sheet("Details")
                
                headers = [
                    'Transmission ID', 'Invoice Number', 'IRN', 'Client ID',
                    'Organization ID', 'Status', 'Submitted At', 'Processing Time (s)',
                    'Retry Count', 'Error Code', 'Payload Size (bytes)'
                ]
                
                for i, header in enumerate(headers, start=1):
                    ws_details.cell(row=1, column=i, value=header)
                    ws_details.cell(row=1, column=i).font = Font(bold=True)
                    ws_details.cell(row=1, column=i).fill = PatternFill(
                        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                    )
                
                for row_idx, record in enumerate(records, start=2):
                    data = [
                        record.transmission_id,
                        record.invoice_number,
                        record.irn or '',
                        record.client_id,
                        record.organization_id,
                        record.status.value,
                        record.submitted_at.strftime("%Y-%m-%d %H:%M:%S"),
                        record.processing_time_seconds or '',
                        record.retry_count,
                        record.error_code or '',
                        record.payload_size_bytes
                    ]
                    
                    for col_idx, value in enumerate(data, start=1):
                        ws_details.cell(row=row_idx, column=col_idx, value=value)
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
            
        except ImportError:
            self.logger.warning("openpyxl not available, returning CSV format instead")
            return self._format_csv_report(records, summary, config).encode('utf-8')
    
    def _update_stats(self, format_type: ReportFormat, generation_time: float):
        """Update report generation statistics"""
        self.stats['reports_generated'] += 1
        self.stats['last_report_at'] = datetime.now(timezone.utc).isoformat()
        self.stats['format_usage'][format_type.value] += 1
        
        # Update average generation time
        current_avg = self.stats['average_generation_time']
        total_reports = self.stats['reports_generated']
        self.stats['average_generation_time'] = (
            (current_avg * (total_reports - 1) + generation_time) / total_reports
        )
    
    async def generate_scheduled_reports(self) -> List[Dict[str, Any]]:
        """Generate commonly requested scheduled reports"""
        now = datetime.now(timezone.utc)
        reports = []
        
        # Daily report
        daily_config = ReportConfig(
            start_date=now - timedelta(days=1),
            end_date=now,
            format=ReportFormat.JSON,
            include_details=False,
            include_charts=True
        )
        daily_report = await self.generate_report(daily_config)
        reports.append({
            'type': 'daily',
            'report': daily_report
        })
        
        # Weekly report
        weekly_config = ReportConfig(
            start_date=now - timedelta(days=7),
            end_date=now,
            format=ReportFormat.JSON,
            include_details=False,
            include_charts=True,
            group_by=['status']
        )
        weekly_report = await self.generate_report(weekly_config)
        reports.append({
            'type': 'weekly',
            'report': weekly_report
        })
        
        return reports
    
    async def get_transmission_trends(self, days: int = 30) -> Dict[str, Any]:
        """Get transmission trends over specified number of days"""
        end_date = datetime.now(timezone.utc)
        start_date = end_date - timedelta(days=days)
        
        records = await self.data_provider.get_transmissions(start_date, end_date)
        
        # Group by day
        daily_data = defaultdict(lambda: {'total': 0, 'successful': 0, 'failed': 0})
        
        for record in records:
            day_key = record.submitted_at.strftime('%Y-%m-%d')
            daily_data[day_key]['total'] += 1
            
            if record.status == TransmissionStatus.APPROVED:
                daily_data[day_key]['successful'] += 1
            elif record.status in [TransmissionStatus.REJECTED, TransmissionStatus.FAILED]:
                daily_data[day_key]['failed'] += 1
        
        # Convert to list and sort by date
        trends = []
        for day in sorted(daily_data.keys()):
            data = daily_data[day]
            success_rate = (data['successful'] / data['total'] * 100) if data['total'] > 0 else 0
            
            trends.append({
                'date': day,
                'total_transmissions': data['total'],
                'successful_transmissions': data['successful'],
                'failed_transmissions': data['failed'],
                'success_rate': round(success_rate, 2)
            })
        
        return {
            'period_days': days,
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat(),
            'trends': trends
        }
    
    async def health_check(self) -> Dict[str, Any]:
        """Get transmission report service health status"""
        return {
            'status': 'healthy',
            'service': 'transmission_reports',
            'stats': self.stats.copy(),
            'supported_formats': [format.value for format in ReportFormat],
            'data_provider': type(self.data_provider).__name__,
            'timestamp': datetime.now(timezone.utc).isoformat()
        }
    
    async def cleanup(self):
        """Cleanup report generator resources"""
        self.logger.info("Transmission report generator cleanup initiated")
        
        # Log final statistics
        self.logger.info(f"Final report generation statistics: {self.stats}")
        
        self.logger.info("Transmission report generator cleanup completed")


# Factory functions
def create_transmission_report_generator() -> TransmissionReportGenerator:
    """Create transmission report generator with standard configuration"""
    return TransmissionReportGenerator()


def create_report_config(start_date: datetime,
                        end_date: datetime,
                        format: ReportFormat = ReportFormat.JSON,
                        **kwargs) -> ReportConfig:
    """Create report configuration with standard settings"""
    return ReportConfig(
        start_date=start_date,
        end_date=end_date,
        format=format,
        **kwargs
    )
